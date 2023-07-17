import random
import string
from datetime import datetime, date,timedelta
# import date
import os
# from flask_sqlalchemy import SQLAlchemy
from models import User,app,TypeDefaut,db,Role
from flask import Flask, render_template, request, redirect, url_for, flash,session,jsonify
# from flask_login import login_user
from flask_login import LoginManager,UserMixin,login_user,login_required,logout_user,current_user
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import bcrypt
import glob
from models import User,db,app,TypeDefaut,Role,Service,Fichier,Fichier_charger,Ticket,Corbeille,UserServiceHistory,Transaction,Type
from flask_migrate import Migrate
from decimal import Decimal
import openpyxl
from openpyxl import load_workbook
from sqlalchemy import desc
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
import hashlib
# import mail
from os.path import join, dirname, realpath
from flask_bcrypt import check_password_hash, generate_password_hash,Bcrypt
from openpyxl import load_workbook
import pandas as pd
from sqlalchemy import create_engine
import csv
import numpy as np

import smtplib
from email.mime.text import MIMEText
# from utils import *
from flask_mail import Mail, Message
from email.mime.multipart import MIMEMultipart
from itsdangerous import URLSafeTimedSerializer, SignatureExpired
from flask_paginate import Pagination, get_page_parameter
from datetime import datetime,timedelta,date
import datetime
from sqlalchemy import extract

# Configuration 

login_manager=LoginManager()
login_manager.init_app(app)
login_manager.login_view='login'
migrate = Migrate(app, db)
mail = Mail(app)
bcrypt = Bcrypt()


app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_USERNAME'] = 'snorange2021@gmail.com'
# app.config['MAIL_PASSWORD'] = 'mjlokyqorvlrzqud'
app.config['MAIL_PASSWORD'] = 'nepeeigsxzhbtwgb'
app.config['MAIL_DEBUG'] = True
app.config['MAIL_PORT'] =465
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USE_TLS'] = False
mail = Mail(app)


with app.app_context(): 
    db.create_all()

                                                    ########################################################
                                                    #                     Utilitaire                       #
                                                    ######################################################## 

def send_daily_reminder_email():
    # Calculate the date for the previous day
    previous_day = datetime.now() - timedelta(days=1)
    previous_day = previous_day.date()

    # Query the tickets that are pending and have a resolution date before the previous day
    pending_tickets = Ticket.query.filter(Ticket.defaut == 'NON',Ticket.date_resolution_max < previous_day).all()

    # Group the tickets by agent
    tickets_by_agent = {}
    for ticket in pending_tickets:
        if ticket.evaluateur in tickets_by_agent:
            tickets_by_agent[ticket.evaluateur].append(ticket)
        else:
            tickets_by_agent[ticket.evaluateur] = [ticket]

    # Send the reminder email to each agent with pending tasks
    for agent, tickets in tickets_by_agent.items():

        # Get agent information
        agent_info = User.query.filter(User.nom_abrege == agent).first()
        if agent_info:
            recipient = agent_info.email
            nom_abrege_agent = agent_info.nom_abrege
            login = agent_info.login
            nom = agent_info.nom

            subject = 'Rappel : Tâches en attente dans QUALITE'
            body = f"Bonjour {nom} {nom_abrege_agent}," \
                   f"\nVous avez des tâches en attente de traitement dans QUALITE. Merci de les prendre en charge." \
                   f"\n\nVoici la liste des tâches en attente :"

            for ticket in tickets:
                body += f"\n\n- Libellé de la tâche : {ticket.libelle_service}" \
                        f"\n  Action attendue : {ticket.description}" \
                        f"\n  Date d'imputation : {ticket.enregistre_le}" \
                        f"\n  Date du jour : {previous_day}" \
                        # f"\n  Délai écoulé : {previous_day - ticket.enregistre_le}"

            body += "\n\nCordialement,\nL'équipe QUALITE"

            msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=["diopabubakr79@gmail.com"])
            msg.body = body
            mail.send(msg)

            # Send a copy to the chef de service or chef de département
            if agent_info.role.role == 'Agent':
                chef_service = User.query.filter(User.service.nom == agent_info.service.nom, User.role.role == 'Chef de Service').first()
                if chef_service:
                    msg_cc = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=['diopb4826@gmail.com'])
                    msg_cc.body = body
                    mail.send(msg_cc)

            # elif agent_info.role.role == 'Chef de Service':
            #     chef_departement = User.query.filter(User.departement == agent_info.departement, User.role == 'Chef de département').first()
            #     if chef_departement:
            #         msg_cc = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=['diopb4826@gmail.com'])
            #         msg_cc.body = body
            #         mail.send(msg_cc)

def envoi_agent(user_id, confirm):
    if confirm == "OUI":
        subject = 'Confirmation du défaut'
        body = "Votre N+1 a confirmé le défaut de traitement qui vous a été imputé."
    else:
        subject = 'Contestation du défaut'
        body = "Votre N+1 a contesté le défaut de traitement qui vous a été imputé et l'a soumis à validation."

    msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=[user_id])
    msg.body = body
    mail.send(msg)

def send_validation_reminder_email(defaut):
    subject ="Les enfants je voulais vous avertir KKKKKKKKKKKKKKKKKK"
    msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=[defaut.user_email])
    msg.body = "Veuillez valider votre défaut."
    # Envoyer l'e-mail
    mail.send(msg)

def schedule_validation_reminder_emails():
    defauts = TypeDefaut.get_defauts_to_remind()

    for defaut in defauts:
        send_validation_reminder_email(defaut)
        defaut.set_last_reminder_date(datetime.now().date())

def envoi_n_plus_one(user_id, confirm):
    if confirm == "OUI":
        subject = 'Confirmation du défaut de votre Supérieur(N+1)'
        body = f"Le défaut de traitement imputé à {user_id} a été confirmé."
    else:
        subject = 'Contestation du défaut de votre Supérieur'
        body = f"Le défaut de traitement imputé à {user_id} a été contesté et soumis à validation."

    msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=['diopabubakr79@gmail.com'])
    msg.body = body
    mail.send(msg)


def generate_code():
    dernier_code = get_last_code_from_database() or 0
    # Générer le nouveau code en incrémentant le dernier code utilisé
    nouveau_code = dernier_code + 1
    return nouveau_code


ALLOWED_EXTENSIONS = {'csv', 'xlsx'}
app.config['UPLOAD_FOLDER'] = 'uploads'
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

##################################################################################################################################
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, user_id)


@app.route("/logout")
def logout():
    logout_user()
    return redirect('login')

@app.route('/transactions')
def transactions():
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 10000  # Nombre de lignes par page
    total = Transaction.query.count()
    pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
    transactions_pagination = Transaction.get_transactions(page=page, per_page=per_page)
    return render_template('transactions.html', transactions_pagination=transactions_pagination,pagination=pagination)

                        ########################################################
                        #                     Dashboard                        #
                        ########################################################


@app.route("/sonatel-gmec/menu", methods=('GET', 'POST'))
@login_required
def home():
    print('cureent_user ====>',current_user.service.nom)
    return render_template('pages/menu.html')


                        ########################################################
                        #                     Profil                           #
                        ########################################################


@app.route("/sonatel-gmec/profils", methods=['POST','GET'])
@login_required
def profils():
    role = Role.query.all()
    users_by_role = User.query.filter(User.role_id == current_user.role_id).all()
    user_session = session['login']
    nom_transac = 'profils'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    return render_template('pages/profils.html', role=role, users_by_role=users_by_role)



@app.route("/gestion_profils/<int:id>", methods=['POST','GET'])
@login_required

def gestion_profils(id):
    role = User.query.filter(User.role_id == id).all()

    user_session = session['login']
    nom_transac = f'gestion_profils/{id}'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    return render_template('gestion_profils.html', role=role)



@app.route('/sonatel-gmec/profile_modif/<string:id>', methods=['POST'])
@login_required
def profile_modif(id):
    user = User.query.get(id)
    user_session = session['login']
    nom_transac = f'profile_modif/{id}'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    if not user:
        flash('Utilisateur non trouvé.')
        return redirect(url_for('monprofil'))

    if request.method == 'POST':
        user.email = request.form.get('email')
        user.login = request.form.get('login')
        user.password = request.form.get('password')

        db.session.commit()
        flash('Utilisateur modifié avec succès.', 'success')

    return redirect(url_for('monprofil'))


@app.route("/sonatel-gmec/monprofil")
def monprofil():
    print('les sessions id',session)
    # Exécution de l'insertion dans la table "transaction"
    user_session = session['login']
    nom_transac = 'monprofil'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    user = current_user
    return render_template('pages/monprofil.html',user=user)


                        ########################################################
                        #                     Athentication                    #
                        ########################################################


@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    # print('===========>',current_user,session)
    if current_user.is_authenticated:
        return redirect(url_for('home'))

    if request.method == 'POST':
        login = request.form.get('login')
        password = request.form.get('password')
        print('===========>',session)


        # print('===========>',current_user.nom)

        user = User.query.filter_by(login=login).first()
        if user:
            # Vérifier le mot de passe
            if user.password == password:
                # Stocker les informations utilisateur en session
                session['logged_in'] = True
                session['id_user'] = user.id
                session['login'] = user.login
                session['email'] = user.email
                session['statut'] = user.state
                session['service'] = user.service.nom
                session['matricule'] = user.matricule
                session['nom_abrege'] = user.nom_abrege
                login_user(user)
                return redirect(url_for('home'))
            elif not user.login:
                flash("Le user %s ne se figure pas dans la base",login)
                return redirect(url_for('login'))
            else:
                flash('Mot de passe incorrect', 'error')
                return redirect(url_for('login'))
        else:
            flash(f'Cet utilisateur {login} n\'existe pas pas la base de données', 'error')
            return redirect(url_for('login'))


    return render_template('pages/login.html')



                            ########################################################
                            #                     Ajouter un User                  #
                            ########################################################


@app.route('/add_user', methods=['GET','POST'])
@login_required
def resolution_utilisateurs():
    if request.method == 'POST':
        prenom = request.form['prenom']
        role = int(request.form['roleid'])
        sigle_service = request.form['sigle_service']
        login = request.form['login']
        nom=request.form['nom']
        email=request.form['email']
        nom_abrege = sigle_service + '_' + prenom.replace(' ', '')

        password = "Son@tel2021"
        print('ooooooooo',password)
        user_session = session['login']
        nom_transac = 'add_user'
        transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)

        # Vérifier si l'utilisateur existe déjà dans la base de données
        existing_user = User.query.filter_by(login=login).first()
        if existing_user:
            flash("L'utilisateur avec le login {} existe déjà.".format(login))

        # Générer le hash du mot de passe
        role = Role.query.get(role)  # Récupérer l'instance de la classe Role avec l'ID de rôle

        user = User(matricule=request.form['matricule'], login=request.form['login'], prenom=request.form['prenom'], nom=nom, role=role,
                    sigle_service=request.form['sigle_service'], service_id=int(request.form['service_id']), state=request.form['statut'], email=email, nom_abrege=nom_abrege, date_debut=datetime.datetime.now(), password="Son@tel2021")

        db.session.add(user)
        db.session.commit()
        pwd = "Son@tel2021"
        subject = 'Notification de la Création de Compte'
        body = f'Bonjour {prenom} {nom} \nVotre Compte a été crée avec succés avec comme mot de passe par défaut {pwd}\n\nCordialement,\nEquipe Qualité'

        msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=[email])
        msg.body = body

        mail.send(msg)



    return render_template('pages/menu.html')


@app.route('/confirm_email/<token>')
def confirm_email(token):
    try:
        email = s.loads(token, salt='email-confirm', max_age=3600)
    except SignatureExpired:
        return flash('<h1>The token is expired!</h1>')
    return redirect(url_for('login'))


                            ########################################################
                            #                     Modif PWD                        #
                            ########################################################


@app.route('/changepassword', methods=['GET', 'POST'])
def changepassword():
    if request.method == 'POST':
        login = request.form['login']
        ancien_mot_de_passe = request.form['ancien']
        nouveau_mot_de_passe = request.form['new']
        confirmer_mot_de_passe = request.form['conf']

        # Récupérer l'utilisateur courant
        utilisateur = User.query.filter(User.login == login).first()

        print('=============>',utilisateur)
        user_session = session['login']
        nom_transac = 'changepassword'
        transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
        db.session.add(transaction)
        db.session.commit()

        # Vérifier si l'ancien mot de passe est correct

        if utilisateur:
            if utilisateur.password != ancien_mot_de_passe:
                flash('Ancien mot de passe incorrect', 'danger')
                return redirect(url_for('changepassword'))

            if len(nouveau_mot_de_passe) > 5:
                # Mettre à jour le mot de passe de l'utilisateur
                utilisateur.password = nouveau_mot_de_passe
                # Vérifier la confirmation du nouveau mot de passe
                if nouveau_mot_de_passe == confirmer_mot_de_passe:
                    db.session.commit()
                    flash('Mot de passe modifié avec succès', 'success')
                    return redirect(url_for('changepassword'))
            else:
                flash('Le mot de passe doit comporter au moins 6 caractères', 'warning')
                return redirect(url_for('changepassword'))

        else:
            flash(f"L'utilisateur {login} n'existe pas dans la base")

    return render_template('pages/login.html')




                                ########################################################
                                #                     Editer PWD                       #
                                ########################################################


@app.route('/modifier_utilisateur/<int:user_id>', methods=['GET', 'POST'])
@login_required
def modifier_utilisateur(user_id):
    user = User.query.get(user_id)
    if not user:
        flash('Utilisateur non trouvé.', 'danger')
        return redirect(url_for('users'))

    if request.method == 'POST':
        user.prenom = request.form.get('prenom')
        user.nom = request.form.get('nom')
        user.role = Role.query.get(int(request.form.get('role')))
        user.service = Service.query.get(int(request.form.get('service')))
        user.sigle_service = request.form.get('sigle_service')
        user.matricule = request.form.get('matricule')
        user.state= request.form.get('statut')
        user.nom_abrege = user.sigle_service+'_'+user.prenom
        if user.state == "Clocturé":
            user.date_fin = datetime.now()

        user_session = session['login']
        nom_transac = 'modifier_utilisateur'
        transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
        db.session.add(transaction)
        db.session.commit()
        flash('Utilisateur modifié avec succès.', 'success')
        return redirect(url_for('users'))

    return render_template('pages/utilisateurs.html', user=user)


@app.route('/delete/<string:user_id>', methods=['GET', 'POST'])
@login_required
def delete(user_id):
    user = User.query.get(user_id)
    corebeille = Corbeille.query.get(user_id)

    if user:
        corebeille = Corbeille(
            matricule=user.matricule,
            login=user.login,
            prenom=user.prenom,
            nom=user.nom,
            role = user.role,
            sigle_service=user.sigle_service,
            service_id=user.service.id,
            role_id=user.role.id,
            state="Clocturé",
            email=user.email,
            nom_abrege=user.nom_abrege,
            date_debut=user.date_debut,
            date_fin=datetime.now(),
            password=user.password
        )
        db.session.add(corebeille)
        db.session.delete(user)

        user_session = session['login']
        nom_transac = 'delete_user'
        transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
        db.session.add(transaction)
        db.session.commit()
        flash('Utilisateur supprimé avec succès.', 'success')
    else:
        flash('Utilisateur introuvable.', 'error')

    return render_template('pages/utilisateurs.html')

@app.route('/delete_dans_corbeille/<int:user_id>')
def delete_dans_cor(user_id):
    corebeille = Corbeille.query.get(user_id)

    if corebeille:
        db.session.delete(corebeille)
        db.session.commit()
    return render_template('corbeille.html')

@app.route('/historique_user')
def historique_user():
    # page = request.args.get(get_page_parameter(), type=int, default=1)
    # per_page = 3  # Nombre de lignes par page

    # # Obtenir la liste paginée des utilisateurs
    # total = UserServiceHistory.query.count()
    # pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
    # user_historique = UserServiceHistory.query.order_by(desc(UserServiceHistory.id)).paginate(page=page, per_page=per_page)
    user_historique = UserServiceHistory.query.order_by(desc(UserServiceHistory.id)).all()
    return render_template('historique.html',user_historique=user_historique)

@app.route('/corbeille')
def get_corbeille():
    user_session = session['login']
    nom_transac = 'corbeille'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    # page = request.args.get(get_page_parameter(), type=int, default=1)
    # per_page = 10  # Nombre de lignes par page
    # # Obtenir la liste paginée des utilisateurs
    # total = Corbeille.query.count()
    # pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
    corbeille_pagination = Corbeille.query.join(Role).join(Service).all()
    return render_template('corbeille.html',corbeille_pagination=corbeille_pagination)



@app.route("/sonatel-sovar/guide-utilisateur", methods=['POST','GET'])
def guide():
    return render_template('pages/faq.html') 



                                ########################################################
                                #                     service                          #
                                ########################################################


@app.route("/sonatel-gmec/services", methods=['POST','GET'])
@login_required
def services():
    dim = date.today() 
    date_saisi= dim.strftime('%d-%m-%Y')
    services = Service.query.all()
    user = User.query.all()
    user_session = session['login']
    nom_transac = 'services'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    return render_template('pages/services.html',services=services,user=user) 


@app.route("/sonatel-gmec/utilisateurs", methods=['POST','GET'])
@login_required
def users():
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 5  # Nombre de lignes par page
    user_session = session['login']
    nom_transac = 'utilisateurs'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    # Obtenir la liste paginée des utilisateurs

    # total = User.query.count()
    # pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
    users_pagination = User.query.join(Role).join(Service).all()
    return render_template('pages/utilisateurs.html', users=users,users_pagination=users_pagination) 



@app.route('/sonatel-gmec/service_users')
@login_required
def service_users():
    # Faire correspondre le service user et le service en cours
    users_service = User.query.filter_by(service=current_user.service).all()
    users_services=User.query.all()
    # flash('blabla','success')
    user_session = session['login']
    nom_transac = f'service_users-{users_service.service.nom}'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    return render_template('service_users.html', users_service=users_service,users_services=users_services)


@app.route('/consulter_services/<int:id>')
@login_required
def consulter_services(id):
    # service = Service.query.get(id)
    users_service = User.query.filter(User.service_id == id).all()
    # consulte = Service.query.filter_by(id=service).first()
    user_session = session['login']
    nom_transac = f'consulter_services/{id}'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    return render_template('service_users.html', users_service=users_service)



                                ########################################################
                                #                     Type - Defaut                    #
                                ########################################################


########################### Premier Version des parametres 
@app.route('/parametrage-defauts', methods=['GET', 'POST'])
@login_required
def parametrage_defauts():

    user_session = session['login']
    nom_transac = 'parametrage-defauts'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    # print('OOOOOOOOOOOOOOOOOOOO',user_session)
    user = User.query.all()
    if request.method == 'POST':
        if TypeDefaut.query.count() == 0:
            print('OOJHHGGGGGHHHH++++>')
            code = "code_124"
        else:
            print("CCCCCCCCCCCCCCCCCCCCCC")
            code = TypeDefaut.get_next_code()

        user_session = session['_user_id']
        type_defaut = request.form.get('type_defaut')
        description_defaut = request.form.get('description_defaut')
        # confirm = request.form.get('oui')
        date_debut = request.form.get('date_debut')
        date_fin = None  # La date de fin est initialisée à None
        commentaires = request.form.get('commentaires')
        validation = request.form.get('validation')
        email_concerne = request.form.get('email')
        evaluer = request.form.get('evaluer')
        n1 = request.form.get('n1')

        # print('confirmation >>>>>',confirm)
        listes = []
        for user in user:
            if email_concerne == user.email:
                user_email = user.email
                service= user.service.nom
                # listes.append({'code':code, 'description':description_defaut,'type_default':type_defaut,'commentaires':commentaires,'validation':validation,'service':service})
                # print("PPPPPPPPPPPPPPPPPPPPP==> ",listes)

            # else:
            #     flash(f"Le mail {email_concerne} est insdisponible")

        if type_defaut == "" or description_defaut == "" or date_debut == None:
            flash('Les champs ne doivent pas être vides', 'danger')
            return redirect(url_for('parametrage_defauts'))

        print("Insertion passé avec succés: com et com_n+1",evaluer,n1)

        new_defaut = TypeDefaut(
            code = code,
            type_defaut=type_defaut,
            description_defaut=description_defaut,
            confirm="NON",
            date_debut=date_debut,
            date_fin=date_fin,
            user_email=user_email,
            commentaires=commentaires,
            validation=validation,
            service=service,
            commentaires_evaluer=evaluer,
            commentaires_n1=n1
        )

        # Ajouter le nouveau défaut à la base de données blabla
        db.session.add(new_defaut)
        db.session.commit()
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 5  # Nombre de lignes par page
    total = TypeDefaut.query.count()
    pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
    defauts = TypeDefaut.query.order_by(desc(TypeDefaut.id)).all()
    elements = TypeDefaut.query.filter(TypeDefaut.service == current_user.service.nom).distinct().order_by(TypeDefaut.date_debut.desc()).all()

    return render_template('parametrage_defauts.html', defauts=defauts,user=user,elements=elements,pagination=pagination)
##################################################################################################################################

@app.route('/modifier-defaut/<int:defaut_id>', methods=['GET', 'POST'])
@login_required
def modifier_defaut(defaut_id):
    
    defaut = TypeDefaut.query.get(defaut_id)
    user_email = User.query.filter_by(email=defaut.user_email).first()
    print('4444444444444',user_email)
    user_session = session['login']
    nom_transac = f'modifier-defaut/{defaut_id}'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    if request.method == 'POST':
        confirm = request.form.get('OUI')

        print(">>>>>>>>>>", confirm)
        if confirm:
            defo = TypeDefaut.query.filter(TypeDefaut.user_email==user_email.email,TypeDefaut.date_debut==defaut.date_debut).all()
            print('fffffffffffffffffffff',[defo.validation for defo in defo])


            
            if "Invalide" in [defo.validation for defo in defo]:

                flash('L\'utilisateur a un defaut qui est à l\'etat invalide, \n Un mail lui sera transmis en guise de validation')
                print("ooooooooooooooooooooooooooooooooooooooooooooooo")
                subject = 'Appel à validation'
                body = f"Votre N+1 vient de faire une tentive de confirmation du défaut de traitement imputé à {defaut.user_email}\n Veuillez valider le defaut qu'on vous a imputé" 
                msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=["diopabubakr79@gmail.com"])
                msg.body = body
                mail.send(msg)
            else:
                if confirm:
                    defauts = TypeDefaut.query.filter_by(user_email=user_email.email,date_debut=defaut.date_debut).all()
                    for defaut in defauts:
                        defaut.confirm = confirm
                    db.session.commit()

                if confirm == "OUI":
                    subject = 'Confirmation de votre N+1'
                    body = f"Votre N+1 vient de confirmer le défaut de traitement imputé à {defaut.user_email}"
                else:
                    subject = 'Contestation des défauts'
                    body = f"Le défaut de traitement imputé à {defaut.user_email} a été contesté et soumis à votre N+1 pour confirmation."

                    # send_validation_reminder_email(defauts)
                    schedule_validation_reminder_emails()

                msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=["diopabubakr79@gmail.com"])
                msg.body = body
                mail.send(msg)

                # On notifie lagent et son N+1
                # envoi_agent(defaut.user_email, confirm)
                # Envoie mail à l'agent
                envoi_agent('diopb4826@gmail.com', confirm)
                # Envoie mail à l'agent N+1
                envoi_n_plus_one(defaut.user_email, confirm)
            

        # Je voudrais separer pour que ne modifier que la confirmation (OUI/NON)
        elif request.form.get('description_defaut') or request.form.get('date_fin') or request.form.get('type_defaut') or request.form.get('commentaires') or request.form.get('validation') or request.form.get('evaluer') or request.form.get('n1'):
            defaut.description_defaut = request.form.get('description_defaut')
            defaut.date_fin = request.form.get('date_fin')
            defaut.type_defaut = request.form.get('type_defaut')
            defaut.commentaires = request.form.get('commentaires')
            defaut.validation = request.form.get('validation')
            evaluer = request.form.get('evaluer')
            n1 = request.form.get('n1')
            # print('===>>>>>',session.get(TypeDefaut, defaut_id),'Object: ',defaut)
            print('===>>>>>',session.get(TypeDefaut, defaut_id))

            defaut.commentaires_evaluer = evaluer
            defaut.commentaires_n1 = n1

            db.session.commit()
        # flash('La description du défaut a été modifiée avec succès!', 'success')
        return redirect(url_for('parametrage_defauts'))

    return render_template('modifier_defaut.html', defaut=defaut)


# Interface de chargement des fichiers plats « Défauts » qui permet de charger les tickets dans la table Fichier c'est à dire le ficheir à charger
@app.route('/chargement-defauts', methods=['GET', 'POST'])
@login_required
def chargement_defauts():
    fichiers = Fichier.query.all()
    user_session = session['login']
    nom_transac = 'chargement-defauts'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    # print(fichiers)
    if request.method == 'POST':
        if 'file' not in request.files:
            # flash('Aucun fichier sélectionné.', 'error')
            return redirect(url_for('chargement_defauts'))
        # On recuere le fichier à partir du formulaire
        file = request.files['file']

        print('============>', file.filename,current_user.id)

        if file.filename == '':
            flash('Aucun fichier sélectionné.', 'error')
            return redirect(url_for('chargement_defauts'))
        
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            print('==========>',file_path)
            file.save(file_path)

            fiichier = Fichier_charger(file.filename,current_user.id)
            db.session.add(fiichier)
            print('==========<',current_user.id)

            fichier = Fichier_charger.query.filter_by(nom=file.filename).first()
            user = session['_user_id']
            print('======GGGGGGG===================>',user)
            if fichier and user == current_user.id:
                if fichier:
                    flash(f'Le fichier {fichier.nom} est déja dans la base', 'success')
            else:

                if os.path.exists(file_path):
                    df = pd.read_excel(file_path)
                    print('==========>',df[['Libellé du Service (complet)', 'TYPE_ECHANT', 'Défaut (OUI/NON)']])

                    # Remplacer les valeurs 'nan' par une valeur par défaut ou une chaîne vide
                    df.replace({np.nan: None}, inplace=True)
                    # Itérer sur les lignes du dataframe et enregistrer dans la base de données
                    for _, row in df.iterrows():
                        file = Fichier(
                                numero_demande                 = row['N° Commande'],
                                enregistre_le                  = row['Enregistré le'],         
                                date_resolution                = row['Date de résolution'],            
                                libelle_service                = row['Libellé du Service (complet)'],
                                saisi_par                      = row['Saisi par'],
                                demandeur                      = row['Demandeur'],
                                demandeur_entite               = row['Demandeur : Entité (complète)'],
                                localisation                   = row['Localisation (complète)'],
                                urgence_utilisateur            = row['Urgence utilisateur'],
                                impact                         = row['Impact'],
                                priorite                       = row['Priorité'],
                                statut_demande                 = row['Statut de la demande'],
                                delai_resolution_hhmm          = row['Délai de résolution (hh:mm)'],
                                delai_resolution_min           = row['Délai de résolution (min)'],
                                resolution_immediate           = row['Resolution immédiate'],
                                resolu_par_groupe              = row['Résolu par (groupe)'],
                                origine_demande                = row['Date de résolution maximum'],
                                date_resolution_maximum        = row['Description'],
                                description                    = row['Résolu par (intervenant)'],
                                resolu_par_intervenant         = row['Résolu par (intervenant)'],
                                service_retard_hhmm            = row['Service : Retard (hh:mm)'],
                                service_retard_min             = row['Service : Retard (min)'],
                                group_fr                       = row['GROUP_FR'],
                                resolution                     = row['Résolution'],
                                sla                            = row['SLA'],
                                beneficiaire_courriel          = row['Bénéficiaire : Courriel'],
                                xa_date_fin_de_mois            = row['XA_DATE_FIN_DE_MOIS'],
                                xb_periode                     = row['XB_PERIODE'],
                                xc_statut_trait                = row['XC_STATUT_TRAIT'],
                                xx_num_sequence                = row['XX_NUM_SEQUENCE'],
                                xx_agent_transfert_dsi         = row['XX_AGENT_TRNSFERT_DSI'],
                                xx_agent_responsable           = row['XX_AGENT_RESPONSABLE'],
                                xx_service                     = row['XX_SERVICE'],
                                xx_intervalle_delai_res        = row['XX_INTERVALLE_DELAI_RES'],
                                xx_delai30min                  = row['XX_DELAI30MIN'],
                                xx_delai1h                     = row['XX_DELAI1H'],
                                xx_delai2h                     = row['XX_DELAI2H'],
                                xx_delai1j                     = row['XX_DELAI1J'],
                                xx_delai2j                     = row['XX_DELAI2J'],
                                xx_respect_delais              = row['XX_RESPECT_DELAIS'],
                                xx_retard_en_jours             = row['XX_RETARD_EN_JOURS'],
                                xx_activite                    = row['XX_ACTYIVITE'],
                                xx_a_comptabiliser             = row['XX_A_COMPTABILISER'],
                                xx_application                 = row['XX_APPLICATION'],
                                xx_dep_traitant                = row['XX_DEP_TRAITANT'],
                                xx_direction                   = row['XX_DIRECTION'],
                                xx_agent_refus                 = row['XX_AGENT_REFUS'],
                                numero                         = row['N°'],
                                type_echant                    = row['TYPE_ECHANT'],
                                defaut                         = row['Défaut (OUI/NON)'],
                                type_description_defaut        = row['Type'],
                                description_du_defaut          = row['Description du Défaut'],
                                commentaires                   = row['Commentaires'],
                                note_defaut                    = row['NOTE_DEFAUT'],
                                agent_escalade                 = row['Agent ESCALADE'],
                                pertinence_escalade            = row['Pertinence ESCALADE'],
                                type_erreur_escalade           = row['TypeErreurEsacalade'],
                                actions_correctives_preventives= row['Actions Correctives/Préventives']
                        )

                        db.session.add(file)
            db.session.commit()


            
            flash('Chargement des défauts effectué avec succès!', 'success')
            return redirect(url_for('chargement_defauts'))
        else:
            flash('Type de fichier non autorisé.', 'error')
            return redirect(url_for('chargement_defauts'))
    return render_template('chargement_defauts.html',fichiers=fichiers)

                                                ###################################################################
                                                #########      Modification apportée pour     #####################
                                                ###################################################################

############################ Parametre de la nouvelle Version 
@app.route("/param_defauts")
@login_required
def param_defauts():
    # Obtenez le mois en cours au format "YYYY/MM"
    current_month = datetime.datetime.now().strftime("%Y/%m")
    types_defaut = Type.query.all()
    print("Date courante:", current_month)

    types = Type.query.distinct(Type.type_defaut).all()

    if current_month in [fichier.xb_periode for fichier in Fichier.query.filter_by(defaut="OUI").all()]:
        tickets = Fichier.query.filter(Fichier.defaut == "OUI", Fichier.xb_periode == current_month).all()
    else:
        # Obtenez le mois précédent au format "YYYY/MM" : ici j'ai recuperer à 2 mois derriére (60 jours), faut le modifier en 30 jours en prod
        previous_month = (datetime.datetime.now() - datetime.timedelta(days=60)).strftime("%Y/%m")
        # Ici j'ai filtré les tickets qui ont un défaut à OUI dont la période correspondant au mois indiqué
        tickets = Fichier.query.filter(Fichier.defaut == "OUI", Fichier.xb_periode == previous_month).all()

    return render_template('param_defauts.html',tickets=tickets,Type=Type,types=types,Fichier=Fichier,types_defaut=types_defaut)

@app.route('/modif_param/<int:id>', methods=['GET', 'POST'])
def modif_param(id):
    tickets = Fichier.query.get(id)
    types = Type.query.get(id)

    

    if request.method == 'POST':
        confirm = request.form.get('OUI')
        if confirm:
            print("Confirmation : " + confirm)
            tickets.confirm = confirm
            # tickets.defaut = "NON"
            db.session.commit()
            if confirm == "OUI":
                subject = 'Confirmation de votre N+1'
                body = f"Votre N+1 vient de confirmer le défaut de traitement imputé à {tickets.demandeur}"
            else:
                subject = 'Contestation des défauts'
                body = f"Le défaut de traitement imputé à {tickets.demandeur} a été contesté et soumis à votre N+1 pour confirmation."

                # send_validation_reminder_email(defauts)
                # schedule_validation_reminder_emails()

            msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=["diopabubakr79@gmail.com"])
            msg.body = body
            mail.send(msg)

            # On notifie lagent et son N+1
            # envoi_agent(defaut.user_email, confirm)
            envoi_agent('diopb4826@gmail.com', confirm)
            envoi_n_plus_one(tickets.demandeur, confirm)
        else:
            tickets.type_description_defaut = request.form.get('type_description_defaut')
            tickets.libelle_service = request.form.get('libelle')
            tickets.saisi_par = request.form.get('saisi')
            tickets.demandeur = request.form.get('demandeur')
            tickets.origine_demande = request.form.get('origine')
            tickets.description = request.form.get('description')
            tickets.resolution = request.form.get('resolution_defaut')
            tickets.type_id = request.form.get('type')
            tickets.confirm = "NON"

            tickets.defaut = request.form.get('defaut')
            tickets.xx_agent_responsable = request.form.get('agent_responsable')
            tickets.xx_agent_refus = request.form.get('agent_refus')
            tickets.type_echeant = request.form.get('type_echeant')
            tickets.description_du_defaut = request.form.get('description_du_defaut')
            tickets.commentaires = request.form.get('commentaires')

            print("Description_defaut :"+request.form.get('description_du_defaut'))
            print("type_description_defaut :"+request.form.get('type_description_defaut'))
            if tickets.description_du_defaut:
                print("SSSSSSSSSSSSSSSSSSSSSS")
                db.session.commit()
            else:
                prin('NOOOOOOOOOOOOOOOOOOOOOOOOOO')

    return redirect(url_for('param_defauts'))



@app.route('/chargement-tickets', methods=['GET', 'POST'])
@login_required
def chargement_tickets():


    user_session = session['login']
    nom_transac = 'chargement-tickets'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    if request.method == 'POST':
        confirmation = request.form.get('confirmation')
        # noconf= request.form.get('noconf')

        if confirmation == 'Oui':
            file = request.files['file']
            liste_utile =[]
            if file.filename == '':
                flash('Aucun fichier sélectionné.', 'error')
                return redirect(url_for('chargement_tickets'))
            
            if file:
                try:
                    df = pd.read_excel(file)

                    # Nombre d'enregistrements rejetés
                    num_rejected = 0  
                    # Liste des enregistrements rejetés
                    rejected_records = []

                    df.fillna('', inplace=True)
                    for index, row in df.iterrows():
                        # print('BBBBBBBBonjour')
                        # Récupérer les valeurs des colonnes
                        if pd.notnull(row['Date de résolution maximum']) or isinstance(row['Date de résolution maximum'], pd.Timestamp):
                            # print('Date de résolution maximum')
                            utilisateur = User.query.filter(User.nom_abrege == row['XX_AGENT_RESPONSABLE'], User.state == 'Actif').first()
                            # print('>>>>>>>>>>>>>>><<<<<<<<<<<<<<<<<<<<<<<<<', utilisateur)
                            if utilisateur:
                                # print('Insertion successful')
                                # Charger l'enregistrement dans la base de données
                                ticket = Ticket(
                                    numero_demande=row['N° Commande'],
                                    enregistre_le=row['Enregistré le'],
                                    date_resolution=row['Date de résolution'],
                                    libelle_service=row['Libellé du Service (complet)'],
                                    demandeur=row['Demandeur'],
                                    statut_demande=row['Statut de la demande'],
                                    resolu_par=row['Résolu par (groupe)'],
                                    origine_demande=row['Origine de la demande'],
                                    date_resolution_max=row['Date de résolution maximum'],
                                    description=row['Description'],
                                    resolution=row['Résolution'],
                                    sla=row['SLA'],
                                    nom_abrege_agent=row['Bénéficiaire : Courriel'],
                                    type_echant=row['TYPE_ECHANT'],
                                    defaut=row['Défaut (OUI/NON)'],
                                    type_defaut=row['Type'],
                                    description_defaut=row['Description du Défaut'],
                                    commentaires_defaut=row['Commentaires'],
                                    periode=row['XB_PERIODE'],
                                    evaluateur=row['XX_AGENT_RESPONSABLE']
                                )
                                
                                db.session.add(ticket)
                            else:
                                # L'enregistrement est rejeté car l'utilisateur n'est pas déclaré ou actif
                                num_rejected += 1
                                rejected_records.append(row.to_dict())
                        else:
                            df.at[index, 'Date de résolution maximum'] = np.nan 

                    # Nombre d'enregistrements chargés avec succès
                    num_loaded = len(df) - num_rejected

                    flash(f"Nombre d'enregistrements à charger: {len(df)}", 'info')
                    flash(f"Nombre d'enregistrements chargés: {num_loaded}", 'success')
                    flash(f"Nombre d'enregistrements rejetés: {num_rejected}", 'warning')
                    # long = len(df)
                    # liste_utile.append(long,num_loaded,num_rejected)
                    
                    if num_rejected > 0:
                        print("PPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPP")
                        folder_path = os.path.join(os.getcwd(), 'files_rejet')
                        os.makedirs(folder_path, exist_ok=True)

                        file_path = os.path.join(folder_path, f'rejected_records_{random.randint(12, 706)+random.randint(2, 6)}_{datetime.datetime.now()}.csv')
                        file_path = os.path.abspath(file_path)

                        field_names = ['N° de demande', 'Enregistré le', 'Date de résolution', 'Libellé du Service (complet)', 'Demandeur', 'Statut de la demande', 'Résolu par (groupe)', 'Origine de la demande', 'Date de résolution maximum', 'Description', 'Résolution', 'SLA', 'Bénéficiaire : Courriel', 'TYPE_ECHANT', 'Défaut (OUI/NON)', 'Type', 'Description du Défaut', 'Commentaires', 'XB_PERIODE', 'XX_AGENT_RESPONSABLE', 'reason', 'row_index']

                        with open(file_path, 'w', newline='') as csv_file:
                            fieldnames = rejected_records[0].keys() 
                            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
                            writer.writeheader()
                            
                            writer.writerows(rejected_records)

                    # Envoyer une mail les concerné
                    data = pd.read_csv(file_path)
                    agents = df['XX_AGENT_RESPONSABLE'].unique()

                    for agent in agents:
                        # Récupérer les informations de l'agent
                        agent_info = User.query.filter(User.nom_abrege == agent).first()
                        if agent_info:
                            recipient = agent_info.email
                            nom_abrege_agent = agent_info.nom_abrege
                            login = agent_info.login
                            nom = agent_info.nom
                            mois = datetime.datetime.now().strftime("%Y/%m")

                            subject = 'Notification de rejet des Tickets'
                            body = f"Bonjour {login} {nom_abrege_agent}," \
                                f"\nLe chargement des échantillons de contrôle des défauts du mois de {mois} est terminé." \
                                f"Nous vous invitons à vous connecter à QUALITE pour traiter/commenter les erreurs critiques vous concernant." \
                                f"\nCordialement," \
                                f"\nL'équipe QUALITE"

                            msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=["diopabubakr79@gmail.com"])
                            msg.body = body
                            mail.send(msg)

                            # Envoyer une copie au N+1
                            # n_plus_one_recipient = agent_info.n_plus_one_email
                            n_plus_one_recipient = "diopb4826@gmail.com"
                            if n_plus_one_recipient:
                                msg = Message(subject, sender=app.config['MAIL_USERNAME'], recipients=[n_plus_one_recipient])
                                msg.body = body
                                mail.send(msg)
                        
                    # send_daily_reminder_email()
                            
                        # flash("E-mails de rejet envoyés avec succès",'success')
                    # db.session.commit()


                    flash("Le chargement des tickets a été validé avec succès.", 'success')
                except Exception as e:
                    flash(f"Une erreur s'est produite lors du chargement du fichier: {str(e)}", 'danger')
            else:
                flash("Aucun fichier sélectionné.", 'error')
        else:
            flash("Le chargement des tickets a été annulé.", 'warning')
        
        return redirect(url_for('chargement_tickets'))
    return render_template('chargement_tickets.html')


@app.route('/sonatel-gmec/details_tickets')
@login_required
def details_tickets():
    
    folder_path = os.path.join(os.getcwd(), 'files_rejet')
    user_session = session['login']
    nom_transac = 'details_tickets'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()

    # Obtenir la liste des fichiers dans le répertoire triés par date de modification
    file_list = glob.glob(os.path.join(folder_path, 'rejected_records_*.csv'))
    file_list.sort(key=os.path.getmtime)

    if not file_list:
        # Gérer le cas où la liste est vide, par exemple, renvoyer une erreur ou un message approprié
        return "Aucun fichier trouvé dans le répertoire"

    latest_file = file_list[-1] 
    df = pd.read_csv(latest_file)
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 10  # Nombre de lignes par page

    total = len(df)
    pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')

    # Obtention des données paginées
    start = (page - 1) * per_page
    end = start + per_page
    paginated_data = df.iloc[start:end]

    return render_template('details_tickets.html', df=df, paginated_data=paginated_data, pagination=pagination)


@app.route('/traitement-ec', methods=['GET', 'POST'])
def traitement_ec():
    if request.method == 'POST':
        # Process defect treatment and validation
        flash('Traitement des EC effectué avec succès!', 'success')
        return redirect(url_for('traitement_ec'))
    return render_template('traitement_ec.html')




                                ########################################################
                                #                          API                         #
                                ########################################################


@app.route('/type_defaut')
@login_required
def type_defaut():
    defaut = Type.query.all()
    user_session = session['login']
    nom_transac = 'type_defaut'
    transaction = Transaction(users_transac=user_session, nom_transac=nom_transac)
    db.session.add(transaction)
    db.session.commit()
    return render_template('typedefaut.html',defaut=defaut)

@app.route('/ajouter_type_defaut', methods=['POST'])
def ajouter_type():
    if request.method == 'POST':
        type_defaut = request.form.get('type_defaut')
        description = request.form.get('description')
        type_defaut = Type(
             type_defaut = type_defaut,
             description = description
        )
        db.session.add(type_defaut)
        db.session.commit()

    return redirect('type_defaut')


@app.route('/api/data')
@login_required
def data():
    query = Fichier.query

    # search filter
    search = request.args.get('search')

    print('====>',request.args)
    if search:
        query = query.filter(db.or_(
            Fichier.demandeur.like(f'%{search}%'),
            Fichier.numero_demande.like(f'%{search}%')
        ))
    total = query.count()

    # sorting
    sort = request.args.get('sort')
    if sort:
        order = []
        for s in sort.split(','):
            direction = s[0]
            name = s[1:]
            if name not in ['numero_commande', 'enregistrer_le', 'date_resolution', 'libelle_service', 'statut_demande','defaut']:
                name = 'numero_commande'
            col = getattr(Fichier, name)
            if direction == '-':
                col = col.desc()
            order.append(col)
        if order:
            query = query.order_by(*order)

    # pagination
    start = request.args.get('start', type=int, default=-1)
    length = request.args.get('length', type=int, default=-1)
    if start != -1 and length != -1:
        query = query.offset(start).limit(length)

    # response
    return {
        'data': [user.to_dict() for user in query],
        'total': total,
    }

@app.route("/vide")
def func():
    Transaction.truncate()
    return redirect(url_for('transactions'))



@app.route("/annuler")
def annuler():
    Transaction.annuler()
    return redirect(url_for('transactions'))


@app.route('/details_params/<int:detail_paramId>')
def vue(detail_paramId):
    
    # role = User.query.filter(User.role_id == id).all()
    types = Type.query.all()
    tickets = Fichier.query.filter_by(defaut="OUI",id=detail_paramId).all()
    return render_template('details_param.html',tickets= tickets,types= types,Type=Type)



# ################################################################## RESOUDRE LE PROBLEME DES SELECT #################################################################

@app.route('/update-type-description', methods=['POST'])
def update_type_description():
    new_value = request.json['newValue']
    return jsonify({'success': True})

@app.route('/update-description', methods=['POST'])
def update_description():
    new_values = request.json['newValues']
    return jsonify({'success': True})




################################################ LANCEMENT DU PROGRAMME

if __name__ == '__main__':
    app.run(debug=True)